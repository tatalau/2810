#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Oct  5 14:43:04 2019

@author: tata
"""
from openpyxl import Workbook
import sqlite3 as sql
#getdata for sqlite database
con = sql.connect('inspections and violations.db')
cur = con.cursor()
getdata = cur.execute("""SELECT violation_code, violation_description, count(*) as count FROM violations GROUP by violation_code;
""").fetchall()

#create new excel
wb = Workbook()
ws =  wb.active
ws.title="ViolationType"
ws['A1'] = "Code"
ws['B1'] = "Description"
ws['C1'] = "Count"

## insert data to excel
for i in getdata:
    ws.append(i)
wb.save("ViolationType.xlsx")
print('Done')