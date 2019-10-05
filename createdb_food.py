#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Sep  9 15:13:03 2019

@author: tata
"""
import openpyxl
import sqlite3
connection = sqlite3.connect("inspections and violations.db")

## Open the two tables provided by itself
workbook = openpyxl.load_workbook("inspections.xlsx")
workbook1 = openpyxl.load_workbook("violations.xlsx")
sheet = workbook["inspections"]
sheet1 = workbook1["violations"]

cursor = connection.cursor()


## create new table inspections with data type
sqlcommond="""create table if not exists inspections(
        activity_data DATE,
        employee_id CHAR(9),
        facility_address VARCHAR(50),
        facility_city VARCHAR(20),
        facility_id CHAR(9),
        facility_name VARCHAR(50),
        facility_state CHAR(2),
        facility_zip VARCHAR(10),
        grade CHAR(1),
        owner_id CHAR(9),
        owner_name VARCHAR(30),
        pe_description VARCHAR(50),
        program_element_pe INTEGER(4),
        program_name VARCHAR(30),
        program_status VARCHAR(8),
        record_id CHAR(9),
        score INTEGER(2),
        serial_number VARCHAR(30) not null primary key,
        service_code INTEGER(5),
        service_description VARCHAR(30)
                
        
        
        )"""


cursor.execute(sqlcommond);

## create new table violations with data type
sqlcommond1="""create table if not exists violations(
        points INTEGER(1),
        serial_number CHAR(9),
        violation_code CHAR(4),
        violation_description VARCHAR(50),
        violation_status VARCHAR(30),
        FOREIGN KEY (serial_number) REFERENCES inspections(serial_number)
        

        )"""


cursor.execute(sqlcommond1);

##get data from excel and insert data in database, get all data. 
for i in range (2,sheet.max_row):
    activity_date = sheet.cell(i,1).value
    employee_id = sheet.cell(i,2).value
    facility_address = sheet.cell(i,3).value
    facility_city = sheet.cell(i,4).value
    facility_id = sheet.cell(i,5).value
    facility_name = sheet.cell(i,6).value
    facility_state = sheet.cell(i,7).value
    facility_zip = sheet.cell(i,8).value
    grade = sheet.cell(i,9).value
    owner_id = sheet.cell(i,10).value
    owner_name = sheet.cell(i,11).value
    pe_description = sheet.cell(i,12).value
    program_element_pe = sheet.cell(i,13).value
    program_name = sheet.cell(i,14).value
    program_status = sheet.cell(i,15).value
    record_id = sheet.cell(i,16).value
    score = sheet.cell(i,17).value
    serial_number = sheet.cell(i,18).value
    service_code = sheet.cell(i,19).value
    service_description = sheet.cell(i,20).value
    insert =""" INSERT INTO inspections VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
    value=(activity_date,employee_id,facility_address,facility_city,facility_id,facility_name,facility_state,facility_zip,grade,owner_id,owner_name,pe_description,program_element_pe,program_name,program_status,record_id,score,serial_number,service_code,service_description)
    cursor.execute(insert,value)

##get data from the excel and insert the data into database, get all data
for a in range (2, sheet1.max_row):
    points = sheet1.cell(a,1).value
    serial_number = sheet1.cell(a,2).value
    violation_code = sheet1.cell(a,3).value
    violation_description = sheet1.cell(a,4).value
    violation_status = sheet1.cell(a,5).value
    insert1 =""" INSERT INTO violations VALUES(?,?,?,?,?)"""
    value1 = (points,serial_number,violation_code,violation_description,violation_status)
    cursor.execute(insert1,value1)

connection.commit()
connection.close()



    